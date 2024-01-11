from ast import pattern
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import  BarChart, Reference



# lê pasta de trabalho e planilha

wb= load_workbook("barcharte.xlsx")
sheet = wb["Relatorio"]

#Referencas das linhas e colunas 

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row


#incluido formulas

#simples

# sheet['B6'] = '=SUM(B2:B5)'
# sheet['B6'].style = "Currency"



for i in range(min_column+1, max_column+1):
    latter = get_column_letter(i)
    sheet[f"{latter}{max_row+1}"] = f"=SUM({latter}{min_row+1}:{latter}{max_row})"
    sheet[f"{latter}{max_row+1}"].style = 'Currency'
    
    
    
wb.save('test.xlsx')